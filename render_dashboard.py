"""
销售业务健康度分析仪表盘 - 数据渲染脚本

功能：
1. 读取Excel销售回款数据
2. 计算所有业务指标（回款率、认购转签约率、来访转认购率等）
3. 渲染HTML仪表盘模板

用法：
    python render_dashboard.py --input 销售回款数据.xlsx --output dashboard.html
"""

import argparse
import json
import openpyxl
from pathlib import Path


def process_excel(filepath: str) -> dict:
    """
    从Excel读取数据并计算所有指标。

    Excel列结构（预期）：
    col1: 项目名称
    col2: 日期
    col3: 点击量_月累计
    col4: 点击量_年累计
    col5: 点击量_年累计_同比增长率
    col6: 来访人次_月累计
    col7: 来访人次_年累计
    col8: 来访人次_年累计_同比增长率
    col9: 点击-来访转化率
    col11: 认购套数_月累计
    col12: 认购套数_年累计
    col13: 认购套数_年累计_去年同期
    col14: 来访转认购率_年累计
    col15: 来访转认购率_去年同期
    col16: 来访转认购转化天数_中位数
    col18: 签约套数_月累计
    col19: 目标签约套数
    col20: 签约套数_年累计
    col21: 目标签约套数_年累计
    col22: 签约套数达成率
    col23: 签约金额_月累计
    col24: 签约金额_年累计
    col25: 目标签约金额
    col26: 目标签约金额_年累计
    col27: 签约金额达成率
    col28: 签约面积_月累计
    col29: 签约面积_年累计
    col31: 目标签约面积_年累计
    col33: 签约均价年累计
    col34: 签约均价年累计_去年同期
    col35: 签约均价_年累计_同比增长率
    col37: 认购转签约率_年累计
    col38: 认购转签约率_去年同期
    col39: 认购转签约天数_中位数
    col41: 应收金额_年累计
    col42: 计划应收金额_年累计
    col43: 实收金额_年累计
    col44: 应收未回款金额_年累计
    """

    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    # 收集所有数据
    projects_data = {}
    all_months = set()

    for r in range(5, ws.max_row + 1):
        proj = ws.cell(row=r, column=1).value
        dt = ws.cell(row=r, column=2).value
        if not proj or not dt:
            continue
        if proj not in projects_data:
            projects_data[proj] = {}
        projects_data[proj][dt] = {
            'clicks_mc': ws.cell(row=r, column=3).value or 0,
            'clicks_yc': ws.cell(row=r, column=4).value or 0,
            'clicks_yoy': ws.cell(row=r, column=5).value or 0,
            'visits_mc': ws.cell(row=r, column=6).value or 0,
            'visits_yc': ws.cell(row=r, column=7).value or 0,
            'visits_yoy': ws.cell(row=r, column=8).value or 0,
            'click_visit_rate': ws.cell(row=r, column=9).value or 0,
            'rg_units_mc': ws.cell(row=r, column=11).value or 0,
            'rg_units_yc': ws.cell(row=r, column=12).value or 0,
            'rg_units_ly': ws.cell(row=r, column=13).value or 0,
            'vi_rg_rate': ws.cell(row=r, column=14).value or 0,
            'vi_rg_rate_ly': ws.cell(row=r, column=15).value or 0,
            'vi_rg_days': ws.cell(row=r, column=16).value or 0,
            'qy_units_mc': ws.cell(row=r, column=18).value or 0,
            'tgt_qy_units': ws.cell(row=r, column=19).value or 0,
            'qy_units_yc': ws.cell(row=r, column=20).value or 0,
            'tgt_qy_units_yc': ws.cell(row=r, column=21).value or 0,
            'qy_units_ach': ws.cell(row=r, column=22).value or 0,
            'qy_amount_mc': ws.cell(row=r, column=23).value or 0,
            'qy_amount_yc': ws.cell(row=r, column=24).value or 0,
            'tgt_qy_amount': ws.cell(row=r, column=25).value or 0,
            'tgt_qy_amount_yc': ws.cell(row=r, column=26).value or 0,
            'qy_amount_ach': ws.cell(row=r, column=27).value or 0,
            'area_mc': ws.cell(row=r, column=28).value or 0,
            'area_yc': ws.cell(row=r, column=29).value or 0,
            'tgt_area_yc': ws.cell(row=r, column=31).value or 0,
            'price_yc': ws.cell(row=r, column=33).value or 0,
            'price_ly': ws.cell(row=r, column=34).value or 0,
            'price_yoy': ws.cell(row=r, column=35).value or 0,
            'rg_qy_rate': ws.cell(row=r, column=37).value or 0,
            'rg_qy_rate_ly': ws.cell(row=r, column=38).value or 0,
            'rg_qy_days': ws.cell(row=r, column=39).value or 0,
            'ys_amount': ws.cell(row=r, column=41).value or 0,
            'plan_ys': ws.cell(row=r, column=42).value or 0,
            'sh_amount': ws.cell(row=r, column=43).value or 0,
            'unpaid': ws.cell(row=r, column=44).value or 0,
        }
        all_months.add(dt)

    # 获取Q1 2026月份
    q1_2026 = sorted([m for m in all_months if m.year == 2026 and m.month <= 3])
    q1_2025 = sorted([m for m in all_months if m.year == 2025 and m.month <= 3])

    # 为每个项目取Q1 2026最新月数据
    projects = {}
    for proj, months in projects_data.items():
        latest_2026 = None
        for m in q1_2026:
            if m in months:
                latest_2026 = months[m]
        if latest_2026 is None:
            continue

        latest_2025 = None
        for m in q1_2025:
            if m in months:
                latest_2025 = months[m]

        # 月度数据
        monthly = []
        for m in q1_2026:
            if m in months:
                d = months[m]
                monthly.append({
                    'date': m.strftime('%Y-%m'),
                    'qy': d['qy_amount_mc'],
                    'sh': d['sh_amount'],
                    'qy_units': d['qy_units_mc'],
                    'rg_units': d['rg_units_mc'],
                    'visits': d['visits_mc'],
                    'tgt': d['tgt_qy_amount'],
                    'unpaid': d['unpaid'],
                })

        d = latest_2026

        # 回款率 = 实收金额 / 应收金额
        ys = d['ys_amount']
        sh = d['sh_amount']
        collection_ratio = sh / ys if ys and ys != 0 else 0

        # 应收未回款率
        unpaid_rate = d['unpaid'] / ys if ys and ys != 0 else 0

        # 实收/签约比
        sh_to_qy_ratio = sh / d['qy_amount_yc'] if d['qy_amount_yc'] and d['qy_amount_yc'] != 0 else 0

        projects[proj] = {
            'qy_amount': round(d['qy_amount_yc'], 2),
            'sh_amount': round(d['sh_amount'], 2),
            'ys_amount': round(d['ys_amount'], 2),
            'qy_units': d['qy_units_yc'],
            'rg_units': d['rg_units_yc'],
            'visits': d['visits_yc'],
            'clicks': d['clicks_yc'],
            'area': round(d['area_yc'], 2),
            'price': round(d['price_yc'], 2),
            'tgt_qy_amount': round(d['tgt_qy_amount_yc'], 2),
            'tgt_qy_units': d['tgt_qy_units_yc'],
            'qy_achievement': round(d['qy_amount_ach'], 4),
            'units_achievement': round(d['qy_units_ach'], 4),
            'collection_ratio': round(collection_ratio, 4),
            'rg_to_qy_rate': round(d['rg_qy_rate'], 4),
            'visit_to_rg_rate': round(d['vi_rg_rate'], 4),
            'click_to_visit_rate': round(d['click_visit_rate'], 4),
            'price_yoy': round(d['price_yoy'], 4),
            'visits_yoy': round(d['visits_yoy'], 4),
            'unpaid': round(d['unpaid'], 2),
            'unpaid_rate': round(unpaid_rate, 4),
            'rg_to_qy_days': round(d['rg_qy_days'], 2),
            'visit_to_rg_days': round(d['vi_rg_days'], 2),
            'rg_to_qy_rate_ly': round(d['rg_qy_rate_ly'], 4),
            'visit_to_rg_rate_ly': round(d['vi_rg_rate_ly'], 4),
            'sh_to_qy_ratio': round(sh_to_qy_ratio, 4),
        }

        # 计算visits_yoy如果Excel没有
        if projects[proj]['visits_yoy'] == 0 and latest_2025 and latest_2025['visits_yc']:
            v2025 = latest_2025['visits_yc']
            if v2025:
                projects[proj]['visits_yoy'] = round(
                    (d['visits_yc'] - v2025) / v2025, 4
                )

    # 计算总计
    total_qy = sum(p['qy_amount'] for p in projects.values())
    total_sh = sum(p['sh_amount'] for p in projects.values())
    total_ys = sum(p['ys_amount'] for p in projects.values())
    total_qy_units = sum(p['qy_units'] for p in projects.values())
    total_rg_units = sum(p['rg_units'] for p in projects.values())
    total_visits = sum(p['visits'] for p in projects.values())
    total_clicks = sum(p['clicks'] for p in projects.values())
    total_area = sum(p['area'] for p in projects.values())
    total_unpaid = sum(p['unpaid'] for p in projects.values())
    total_tgt_qy = sum(p['tgt_qy_amount'] for p in projects.values())
    total_tgt_units = sum(p['tgt_qy_units'] for p in projects.values())

    # 回款率 = 实收/应收
    total_coll_ratio = total_sh / total_ys if total_ys and total_ys != 0 else 0
    total_unpaid_rate = total_unpaid / total_ys if total_ys and total_ys != 0 else 0

    # 价格 = 总金额/总面积
    total_price = total_qy / total_area if total_area and total_area != 0 else 0

    # 认购转签约率 = sum(认转签套数)/sum(认购套数)
    total_rg_to_qy_units = sum(p['rg_to_qy_rate'] * p['rg_units'] for p in projects.values())
    total_rg_to_qy_rate = total_rg_to_qy_units / total_rg_units if total_rg_units and total_rg_units != 0 else 0

    # 来访转认购率 = sum(来转认套数)/sum(来访)
    total_vi_rg_units = sum(p['visit_to_rg_rate'] * p['visits'] for p in projects.values())
    total_vi_rg_rate = total_vi_rg_units / total_visits if total_visits and total_visits != 0 else 0

    # 转换天数中位数
    def median(lst):
        if not lst:
            return 0
        s = sorted(lst)
        n = len(s)
        if n % 2 == 1:
            return s[n // 2]
        return (s[n // 2 - 1] + s[n // 2]) / 2

    rg_days_list = [p['rg_to_qy_days'] for p in projects.values() if p['rg_to_qy_days'] > 0]
    vi_days_list = [p['visit_to_rg_days'] for p in projects.values() if p['visit_to_rg_days'] > 0]

    total_rg_days = median(rg_days_list)
    total_vi_days = median(vi_days_list)

    # 实收/签约比
    total_sh_to_qy = total_sh / total_qy if total_qy and total_qy != 0 else 0

    # 去年同期合计
    total_qy_2025 = 0
    total_sh_2025 = 0
    total_units_2025 = 0
    total_visits_2025 = 0
    for proj, months in projects_data.items():
        for m in q1_2025:
            if m in months:
                d = months[m]
                total_qy_2025 += d['qy_amount_yc']
                total_sh_2025 += d['sh_amount']
                total_units_2025 += d['qy_units_yc']
                total_visits_2025 += d['visits_yc']

    # 月度合计
    monthly_totals = []
    for m in q1_2026:
        mq = sum(projects_data[p][m]['qy_amount_mc'] for p in projects_data if m in projects_data[p])
        ms = sum(projects_data[p][m]['sh_amount'] for p in projects_data if m in projects_data[p])
        mu = sum(projects_data[p][m]['qy_units_mc'] for p in projects_data if m in projects_data[p])
        mrg = sum(projects_data[p][m]['rg_units_mc'] for p in projects_data if m in projects_data[p])
        mv = sum(projects_data[p][m]['visits_mc'] for p in projects_data if m in projects_data[p])
        mt = sum(projects_data[p][m]['tgt_qy_amount'] for p in projects_data if m in projects_data[p])
        mup = sum(projects_data[p][m]['unpaid'] for p in projects_data if m in projects_data[p])
        monthly_totals.append({
            'date': m.strftime('%Y-%m'),
            'qy': round(mq, 2),
            'sh': round(ms, 2),
            'qy_units': mu,
            'rg_units': mrg,
            'visits': mv,
            'tgt': round(mt, 2),
            'unpaid': round(mup, 2),
        })

    totals = {
        'qy_amount': round(total_qy, 2),
        'sh_amount': round(total_sh, 2),
        'ys_amount': round(total_ys, 2),
        'qy_units': total_qy_units,
        'rg_units': total_rg_units,
        'visits': total_visits,
        'clicks': total_clicks,
        'area': round(total_area, 2),
        'price': round(total_price, 2),
        'tgt_qy_amount': round(total_tgt_qy, 2),
        'tgt_qy_units': total_tgt_units,
        'unpaid': round(total_unpaid, 2),
        'unpaid_rate': round(total_unpaid_rate, 4),
        'rg_to_qy_days': round(total_rg_days, 2),
        'visit_to_rg_days': round(total_vi_days, 2),
        'rg_to_qy_rate': round(total_rg_to_qy_rate, 4),
        'visit_to_rg_rate': round(total_vi_rg_rate, 4),
        'collection_ratio': round(total_coll_ratio, 4),
        'sh_to_qy_ratio': round(total_sh_to_qy, 4),
        'q1_2025_qy': round(total_qy_2025, 2),
        'q1_2025_sh': round(total_sh_2025, 2),
        'q1_2025_units': total_units_2025,
        'q1_2025_visits': total_visits_2025,
        'monthly': monthly_totals,
    }

    out_projects = []
    for proj, d in sorted(projects.items()):
        p = {
            'project': proj,
            'qy_amount': round(d['qy_amount'], 2),
            'sh_amount': round(d['sh_amount'], 2),
            'ys_amount': round(d['ys_amount'], 2),
            'qy_units': d['qy_units'],
            'rg_units': d['rg_units'],
            'visits': d['visits'],
            'clicks': d['clicks'],
            'area': round(d['area'], 2),
            'price': round(d['price'], 2),
            'tgt_qy_amount': round(d['tgt_qy_amount'], 2),
            'tgt_qy_units': d['tgt_qy_units'],
            'qy_achievement': round(d['qy_achievement'], 4),
            'units_achievement': round(d['units_achievement'], 4),
            'collection_ratio': round(d['collection_ratio'], 4),
            'rg_to_qy_rate': round(d['rg_to_qy_rate'], 4),
            'visit_to_rg_rate': round(d['visit_to_rg_rate'], 4),
            'click_to_visit_rate': round(d['click_to_visit_rate'], 4),
            'price_yoy': round(d['price_yoy'], 4),
            'visits_yoy': round(d['visits_yoy'], 4),
            'unpaid': round(d['unpaid'], 2),
            'unpaid_rate': round(d['unpaid_rate'], 4),
            'rg_to_qy_days': round(d['rg_to_qy_days'], 2),
            'visit_to_rg_days': round(d['visit_to_rg_days'], 2),
            'rg_to_qy_rate_ly': round(d['rg_to_qy_rate_ly'], 4),
            'visit_to_rg_rate_ly': round(d['visit_to_rg_rate_ly'], 4),
            'sh_to_qy_ratio': round(d['sh_to_qy_ratio'], 4),
        }
        out_projects.append(p)

    return {'totals': totals, 'projects': out_projects}


def render_dashboard(template_path: str, data: dict) -> str:
    """
    将计算好的数据注入HTML模板，返回完整的HTML字符串。
    """
    with open(template_path, 'r', encoding='utf-8') as f:
        template = f.read()

    data_json = json.dumps(data, ensure_ascii=False)

    rendered = template.replace('{{DATA_JSON}}', data_json)
    return rendered


def main():
    parser = argparse.ArgumentParser(description='销售业务健康度仪表盘渲染')
    parser.add_argument('--input', '-i', required=True, help='Excel数据文件路径')
    parser.add_argument('--template', '-t', default='dashboard_template.html', help='HTML模板路径')
    parser.add_argument('--output', '-o', default='dashboard_output.html', help='输出HTML路径')
    args = parser.parse_args()

    print(f"📊 读取数据: {args.input}")
    data = process_excel(args.input)
    print(f"   ✅ {len(data['projects'])} 个项目")
    print(f"   签约总额: ¥{data['totals']['qy_amount']/1e4:.0f}万")
    print(f"   实收总额: ¥{data['totals']['sh_amount']/1e4:.0f}万")
    print(f"   回款率: {data['totals']['collection_ratio']*100:.1f}%")

    print(f"🎨 渲染仪表盘: {args.template} → {args.output}")
    html = render_dashboard(args.template, data)
    with open(args.output, 'w', encoding='utf-8') as f:
        f.write(html)

    print(f"✅ 完成! 输出: {args.output}")


if __name__ == '__main__':
    main()
